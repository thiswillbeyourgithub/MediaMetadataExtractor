#!/usr/bin/env python3
"""Media Metadata Extractor

This application extracts metadata from media files including:
- Duration
- Resolution (width x height)
- FPS (frames per second)
- Codec information
- File size
- Creation date (if available in EXIF)

The results are saved to an Excel file with each file's metadata in a row.

Key Features:
- Supports common media formats: .mp3, .mp4, .avi, .mkv, .mov, .wav, .flac, .m4a, .aac
- Recursively scans directories
- Excludes hidden files (those starting with '.')
- Provides:
  - Total number of media files
  - Total size in GB
  - Detailed metadata for each file
  - Results saved to Excel file
  - Remembers last used directory

Dependencies:
- moviepy: For basic media metadata
- openpyxl: For Excel file creation
"""

import os
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from pathlib import Path
from typing import Dict, List
import warnings
from datetime import datetime
import threading
from openpyxl import Workbook
from openpyxl.styles import Font
from moviepy.video.io.VideoFileClip import VideoFileClip

MEDIA_EXTENSIONS = {'.mp3', '.mp4', '.avi', '.mkv', '.mov', '.wav', '.flac', '.m4a', '.aac'}

def get_media_metadata(file_path: Path) -> Dict[str, str]:
    """Extract metadata from a media file.
    
    Args:
        file_path: Path to the media file
        
    Returns:
        Dictionary containing extracted metadata
    """
    metadata = {
        'filename': file_path.name,
        'path': str(file_path),
        'size_B': file_path.stat().st_size,
        'size_MB': f"{file_path.stat().st_size / (1024 * 1024):.2f}",
        'modified_unix': file_path.stat().st_mtime,
        'modified_date': datetime.fromtimestamp(file_path.stat().st_mtime).strftime('%Y-%m-%d %H:%M:%S'),
    }
    
    try:
        # Suppress warnings during metadata extraction
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            with VideoFileClip(str(file_path)) as clip:
                hours = clip.duration//60//60
                minutes = clip.duration//60 - hours * 60
                seconds = clip.duration - hours * 60 * 60 - minutes * 60
                metadata.update({
                    'duration_seconds': f"{clip.duration:.2f}",
                    'duration': f"{hours}H{minutes}::{seconds:.2f}",
                    'resolution': f"{clip.size[0]}x{clip.size[1]}",
                    'color_space': clip.color_space,
                    'fps': f"{clip.fps:.2f}" if hasattr(clip, 'fps') else 'N/A',
                    'codec': clip.reader.codec if hasattr(clip.reader, 'codec') else 'N/A',
                    'pixel_format': clip.reader.pixel_format if hasattr(clip.reader, 'pixel_format') else 'N/A',
                    'depth': clip.reader.depth if hasattr(clip.reader, 'depth') else 'N/A',
                    'rotation': clip.reader.rotation if hasattr(clip.reader, 'rotation') else 'N/A',
                    'bitrate': clip.reader.bitrate if hasattr(clip.reader, 'bitrate') else 'N/A',
                    'extra_infos': clip.reader.infos if hasattr(clip.reader, 'infos') else 'N/A',
                })
    except Exception as e:
        metadata['error'] = str(e)
    
    return metadata

def save_to_excel(data: List[Dict[str, str]], output_path: Path) -> None:
    """Save metadata to an Excel file.
    
    Args:
        data: List of metadata dictionaries
        output_path: Path to save the Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Media Metadata"
    
    # Create header row
    headers = [
        'Filename', 'Path', 'Size (B)', 'Size (MB)', 'Modified Date', 
        'Duration (seconds)', 'Duration', 'Resolution', 'FPS', 
        'Codec', 'Pixel Format', 'Bit Depth', 'Rotation', 
        'Bitrate', 'Color Space', 'Extra Infos'
    ]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
    
    # Add data rows
    for row_num, item in enumerate(data, 2):
        ws.cell(row=row_num, column=1, value=item['filename'])
        ws.cell(row=row_num, column=2, value=item['path'])
        ws.cell(row=row_num, column=3, value=item['size_B'])
        ws.cell(row=row_num, column=4, value=item['size_MB'])
        ws.cell(row=row_num, column=5, value=item['modified_date'])
        ws.cell(row=row_num, column=6, value=item.get('duration_seconds', 'N/A'))
        ws.cell(row=row_num, column=7, value=item.get('duration', 'N/A'))
        ws.cell(row=row_num, column=8, value=item.get('resolution', 'N/A'))
        ws.cell(row=row_num, column=9, value=item.get('fps', 'N/A'))
        ws.cell(row=row_num, column=10, value=item.get('codec', 'N/A'))
        ws.cell(row=row_num, column=11, value=item.get('pixel_format', 'N/A'))
        ws.cell(row=row_num, column=12, value=item.get('depth', 'N/A'))
        ws.cell(row=row_num, column=13, value=item.get('rotation', 'N/A'))
        ws.cell(row=row_num, column=14, value=item.get('bitrate', 'N/A'))
        ws.cell(row=row_num, column=15, value=item.get('color_space', 'N/A'))
        ws.cell(row=row_num, column=16, value=item.get('extra_infos', 'N/A'))
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    wb.save(output_path)

def process_directory(directory: Path, output_file: Path) -> None:
    """Process all media files in a directory and save metadata to Excel.
    
    Args:
        directory: Path to directory containing media files
        output_file: Path to save the Excel file
    """
    if not directory.exists():
        raise FileNotFoundError(f"Directory not found: {directory}")
    
    media_files = [
        f for f in directory.rglob('*') 
        if f.suffix.lower() in MEDIA_EXTENSIONS and not f.name.startswith('.')
    ]
    
    if not media_files:
        raise ValueError(f"No supported media files found in {directory}")
    
    print(f"Found {len(media_files)} media files. Processing...")
    
    metadata_list = []
    for file in media_files:
        print(f"Processing: {file.name}")
        metadata = get_media_metadata(file)
        metadata_list.append(metadata)
    
    print(f"Saving results to {output_file}")
    save_to_excel(metadata_list, output_file)
    print("Processing complete!")

class MediaMetadataExtractor:
    def __init__(self, root):
        self.root = root
        self.root.title("Media Metadata Extractor")
        self.root.geometry("500x400")
        self.root.minsize(400, 300)
        
        # Try to load last used path
        last_path = self._load_last_path()
        self.folder_path = tk.StringVar(value=last_path if last_path else "")
        
        # Main container using grid
        self.main_container = ttk.Frame(root)
        self.main_container.pack(fill="both", expand=True, padx=5, pady=5)
        
        # Documentation section
        self.doc_frame = ttk.Frame(self.main_container)
        self.doc_frame.grid(row=0, column=0, sticky="ew", padx=5, pady=2)
        
        self.doc_label = ttk.Label(
            self.doc_frame, 
            text="▲ Documentation (click to expand)",
            cursor="hand2",
            foreground="blue"
        )
        self.doc_label.pack(fill="x")
        self.doc_label.bind("<Button-1>", self.toggle_documentation)
        
        self.doc_text = tk.Text(
            self.doc_frame, 
            height=5,
            wrap="word", 
            state="disabled",
            padx=5,
            pady=5
        )
        self.doc_text.config(state="normal")
        self.doc_text.insert("1.0", __doc__)
        self.doc_text.config(state="disabled")
        self.docs_visible = False
        
        # Folder selection
        self.folder_frame = ttk.LabelFrame(self.main_container, text="Select Folder")
        self.folder_frame.grid(row=1, column=0, sticky="ew", padx=5, pady=2)
        
        self.folder_entry = ttk.Entry(self.folder_frame, textvariable=self.folder_path)
        self.folder_entry.pack(side="left", fill="x", expand=True, padx=5, pady=5)
        
        self.browse_button = ttk.Button(self.folder_frame, text="Browse", command=self.select_folder)
        self.browse_button.pack(side="right", padx=5, pady=5)
        
        # Output options
        self.output_frame = ttk.LabelFrame(self.main_container, text="Output Options")
        self.output_frame.grid(row=2, column=0, sticky="ew", padx=5, pady=2)
        
        self.output_path = tk.StringVar(value=str(Path.home() / "media_metadata.xlsx"))
        self.output_entry = ttk.Entry(self.output_frame, textvariable=self.output_path)
        self.output_entry.pack(side="left", fill="x", expand=True, padx=(0,5))
        
        self.output_browse_button = ttk.Button(
            self.output_frame, 
            text="Browse", 
            command=self.select_output_file
        )
        self.output_browse_button.pack(side="right")
        
        # Progress
        self.progress_frame = ttk.LabelFrame(self.main_container, text="Progress")
        self.progress_frame.grid(row=3, column=0, sticky="nsew", padx=5, pady=2)
        
        self.progress_text = tk.Text(self.progress_frame, height=8, state="disabled")
        scrollbar = ttk.Scrollbar(self.progress_frame, command=self.progress_text.yview)
        self.progress_text.configure(yscrollcommand=scrollbar.set)
        
        self.progress_text.grid(row=0, column=0, sticky="nsew")
        scrollbar.grid(row=0, column=1, sticky="ns")
        
        self.progress_frame.grid_rowconfigure(0, weight=1)
        self.progress_frame.grid_columnconfigure(0, weight=1)
        self.main_container.grid_rowconfigure(3, weight=1)
        self.main_container.grid_columnconfigure(0, weight=1)
        
        # Control buttons
        self.button_frame = ttk.Frame(self.main_container)
        self.button_frame.grid(row=4, column=0, sticky="ew", padx=5, pady=5)
        
        self.start_button = ttk.Button(
            self.button_frame, 
            text="Start Processing", 
            command=self.start_processing,
            style='Accent.TButton'
        )
        self.start_button.pack(side="left", expand=True, fill="x")
        
        self.cancel_button = ttk.Button(
            self.button_frame,
            text="Cancel",
            command=self.cancel_processing,
            state="disabled"
        )
        self.cancel_button.pack(side="right", expand=True, fill="x")
        
        # Thread control
        self.processing_thread = None
        self.cancel_requested = False
        
        # Configure styles
        style = ttk.Style()
        style.configure('Accent.TButton', 
                       font=('Helvetica', 12, 'bold'),
                       padding=10,
                       foreground='white',
                       background='#0078d7')

    def _get_last_path_file(self) -> Path:
        """Get the path to the last path file in system temp directory."""
        import tempfile
        temp_dir = Path(tempfile.gettempdir())
        return temp_dir / "MediaMetadataExtractor_latest_path.txt"

    def _is_valid_path(self, path: str) -> bool:
        """Check if a path is valid and accessible."""
        try:
            return os.path.exists(path) and os.path.isdir(path)
        except Exception:
            return False

    def _save_last_path(self, path: str) -> None:
        """Save the last selected path to a temporary file only if valid."""
        try:
            if self._is_valid_path(path):
                last_path_file = self._get_last_path_file()
                with open(last_path_file, "w") as f:
                    f.write(path)
        except Exception:
            pass

    def _load_last_path(self) -> Optional[str]:
        """Load the last selected path from temporary file if it exists and is valid."""
        try:
            last_path_file = self._get_last_path_file()
            if last_path_file.exists():
                with open(last_path_file, "r") as f:
                    path = f.read().strip()
                    if path and self._is_valid_path(path):
                        return path
                    os.remove(last_path_file)
        except Exception:
            pass
        return None

    def select_folder(self):
        initialdir = self.folder_path.get() if self.folder_path.get() else None
        folder = filedialog.askdirectory(initialdir=initialdir)
        if folder:
            if self._is_valid_path(folder):
                self.folder_path.set(folder)
                self._save_last_path(folder)
                # Set default output path
                self.output_path.set(str(Path(folder) / "media_metadata.xlsx"))
            else:
                messagebox.showerror(
                    "Invalid Path",
                    f"The selected path is not accessible:\n{folder}\n\n"
                    "Please select a valid directory."
                )

    def select_output_file(self):
        output_file = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile="media_metadata.xlsx"
        )
        if output_file:
            self.output_path.set(output_file)

    def log_message(self, message):
        """Immediate logging for important messages"""
        self.progress_text.config(state="normal")
        self.progress_text.insert("end", message + "\n")
        self.progress_text.see("end")
        self.progress_text.config(state="disabled")

    def start_processing(self):
        folder = self.folder_path.get()
        if not folder:
            messagebox.showerror("Error", "Please select a folder first")
            return
            
        if not os.path.exists(folder):
            messagebox.showerror("Error", f"The selected path does not exist:\n{folder}")
            return
            
        if not os.path.isdir(folder):
            messagebox.showerror("Error", f"The selected path is not a directory:\n{folder}")
            return
            
        self.cancel_requested = False
        self.browse_button.config(state="disabled")
        self.start_button.config(state="disabled")
        self.cancel_button.config(state="normal")
        
        self.processing_thread = threading.Thread(
            target=self.process_folder,
            args=(folder,),
            daemon=True
        )
        self.processing_thread.start()

    def process_folder(self, folder):
        try:
            path = Path(folder)
            media_files = [
                f for f in path.rglob('*') 
                if f.suffix.lower() in MEDIA_EXTENSIONS and not f.name.startswith('.')
            ]
            
            if not media_files:
                messagebox.showinfo("No Files", "No supported media files found in selected directory")
                return
                
            total_size = sum(f.stat().st_size for f in media_files)
            total_size_gb = total_size / (1024 ** 3)
            
            self.log_message(f"Found {len(media_files)} media files ({total_size_gb:.2f} GB)")
            
            metadata_list = []
            for i, file in enumerate(media_files):
                if self.cancel_requested:
                    self.log_message("\nProcessing cancelled by user")
                    break
                    
                self.log_message(f"Processing: {file.name}")
                metadata = get_media_metadata(file)
                metadata_list.append(metadata)
                
                if (i+1) % 10 == 0:  # Update progress every 10 files
                    self.log_message(f"Processed {i+1}/{len(media_files)} files")
            
            output_path = Path(self.output_path.get())
            save_to_excel(metadata_list, output_path)
            self.log_message(f"\nResults saved to {output_path}")
            messagebox.showinfo("Complete", "Metadata extraction finished!")
            
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {str(e)}")
        finally:
            self.browse_button.config(state="normal")
            self.start_button.config(state="normal")
            self.cancel_button.config(state="disabled")
            self.processing_thread = None

    def cancel_processing(self):
        if self.processing_thread and self.processing_thread.is_alive():
            self.cancel_requested = True
            self.log_message("\nCancelling... Please wait for current file to finish.")
            self.cancel_button.config(state="disabled")

    def toggle_documentation(self, event=None):
        if self.docs_visible:
            self.doc_text.pack_forget()
            self.doc_label.config(text="▲ Documentation (click to expand)")
            self.docs_visible = False
        else:
            self.doc_text.pack(fill="x", padx=5, pady=5)
            self.doc_label.config(text="▼ Documentation (click to collapse)")
            self.docs_visible = True

if __name__ == '__main__':
    try:
        root = tk.Tk()
        app = MediaMetadataExtractor(root)
        root.mainloop()
    except KeyboardInterrupt:
        print("\nKeyboard interrupt received. Exiting gracefully...")
        try:
            if app.processing_thread and app.processing_thread.is_alive():
                app.cancel_processing()
                app.processing_thread.join(timeout=2)
        except Exception:
            pass
        root.destroy()
