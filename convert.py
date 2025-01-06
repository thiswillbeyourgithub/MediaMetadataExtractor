import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font
from collections import defaultdict

def group_by_folder(data):
    """Group metadata entries by their folder path and sort by filename"""
    grouped = defaultdict(list)
    for item in data:
        folder = str(Path(item['path']).parent)
        grouped[folder].append(item)
    
    # Sort items in each folder by filename
    for folder in grouped:
        grouped[folder].sort(key=lambda x: x['filename'])
    
    return grouped

def create_sheet(wb, folder, data):
    """Create a worksheet for a folder with its metadata"""
    # Replace / with __ and truncate to 31 chars for Excel sheet name
    sheet_title = folder.replace('/', '__')[:31]
    ws = wb.create_sheet(title=sheet_title)
    
    # Create header row
    headers = [
        'Filename', 'Path', 'Size (B)', 'Size (MB)', 'Modified Date', 
        'Duration (seconds)', 'Duration', 'Resolution', 'FPS', 
        'Codec', 'Pixel Format', 'Bit Depth', 'Rotation', 
        'Bitrate', 'Color Space', 'Extra Infos'
    ]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
    
    # Add data rows
    for row_num, item in enumerate(data, 2):
        ws.cell(row=row_num, column=1, value=item['filename'])
        ws.cell(row=row_num, column=2, value=item['path'])
        ws.cell(row=row_num, column=3, value=item['size_B'])
        ws.cell(row=row_num, column=4, value=item['size_MB'])
        ws.cell(row=row_num, column=5, value=item['modified_date'])
        ws.cell(row=row_num, column=6, value=item.get('duration_seconds', 'N/A'))
        ws.cell(row=row_num, column=7, value=item.get('duration', 'N/A'))
        ws.cell(row=row_num, column=8, value=item.get('resolution', 'N/A'))
        ws.cell(row=row_num, column=9, value=item.get('fps', 'N/A'))
        ws.cell(row=row_num, column=10, value=item.get('codec', 'N/A'))
        ws.cell(row=row_num, column=11, value=item.get('pixel_format', 'N/A'))
        ws.cell(row=row_num, column=12, value=item.get('depth', 'N/A'))
        ws.cell(row=row_num, column=13, value=item.get('rotation', 'N/A'))
        ws.cell(row=row_num, column=14, value=item.get('bitrate', 'N/A'))
        ws.cell(row=row_num, column=15, value=item.get('color_space', 'N/A'))
        ws.cell(row=row_num, column=16, value=item.get('extra_infos', 'N/A'))
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width

def convert_json_to_excel(json_path, output_path):
    """Convert JSON metadata to Excel with sheets per folder"""
    # Load JSON data
    with open(json_path, 'r') as f:
        data = json.load(f)
    
    # Group data by folder
    grouped_data = group_by_folder(data)
    
    # Create Excel workbook
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create sheets for each folder
    for folder, items in grouped_data.items():
        create_sheet(wb, folder, items)
    
    # Save workbook
    wb.save(output_path)
    print(f"Excel file saved to {output_path}")

if __name__ == '__main__':
    import sys
    if len(sys.argv) != 3:
        print("Usage: python convert.py <input.json> <output.xlsx>")
        sys.exit(1)
    
    json_file = sys.argv[1]
    excel_file = sys.argv[2]
    convert_json_to_excel(json_file, excel_file)
